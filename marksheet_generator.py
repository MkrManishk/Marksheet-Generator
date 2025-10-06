import tkinter as tk
from tkinter import ttk, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class MarksheetGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Marksheet Generator")
        self.root.geometry("900x700")
        self.root.configure(bg="#f5f5f5")
        self.root.resizable(True, True)

        self.subject_entries = []
        self.marks_entries = []
        self.total_marks_entries = []

        self.setup_ui()

    def setup_ui(self):
        main_frame = tk.Frame(self.root, bg="#f5f5f5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        title_label = tk.Label(
            main_frame,
            text="Marksheet Generator",
            font=("Segoe UI", 24, "bold"),
            bg="#f5f5f5",
            fg="#2c3e50"
        )
        title_label.pack(pady=(0, 20))

        content_frame = tk.Frame(main_frame, bg="#f5f5f5")
        content_frame.pack(fill=tk.BOTH, expand=True)

        left_frame = tk.Frame(content_frame, bg="#ffffff", relief=tk.FLAT, bd=0)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        left_frame.configure(highlightbackground="#e0e0e0", highlightthickness=1)

        input_container = tk.Frame(left_frame, bg="#ffffff")
        input_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        self.create_student_info_section(input_container)
        self.create_subjects_section(input_container)
        self.create_buttons_section(input_container)
        self.create_results_section(input_container)

        right_frame = tk.Frame(content_frame, bg="#ffffff", relief=tk.FLAT, bd=0)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        right_frame.configure(highlightbackground="#e0e0e0", highlightthickness=1)

        self.create_chart_section(right_frame)

    def create_student_info_section(self, parent):
        info_frame = tk.LabelFrame(
            parent,
            text="Student Information",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        info_frame.pack(fill=tk.X, pady=(0, 15))

        self.name_entry = self.create_input_field(info_frame, "Student Name:", 0)
        self.roll_entry = self.create_input_field(info_frame, "Roll Number:", 1)
        self.class_entry = self.create_input_field(info_frame, "Class/Section:", 2)

    def create_subjects_section(self, parent):
        subjects_frame = tk.LabelFrame(
            parent,
            text="Subject Marks",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        subjects_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        headers_frame = tk.Frame(subjects_frame, bg="#ffffff")
        headers_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(headers_frame, text="Subject", font=("Segoe UI", 9, "bold"),
                bg="#ffffff", fg="#2c3e50", width=15, anchor='w').pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(headers_frame, text="Marks Obtained", font=("Segoe UI", 9, "bold"),
                bg="#ffffff", fg="#2c3e50", width=15, anchor='w').pack(side=tk.LEFT, padx=5)
        tk.Label(headers_frame, text="Total Marks", font=("Segoe UI", 9, "bold"),
                bg="#ffffff", fg="#2c3e50", width=15, anchor='w').pack(side=tk.LEFT, padx=5)

        for i in range(5):
            self.create_subject_row(subjects_frame, i)

    def create_subject_row(self, parent, index):
        row_frame = tk.Frame(parent, bg="#ffffff")
        row_frame.pack(fill=tk.X, pady=3)

        subject_entry = tk.Entry(
            row_frame,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bg="#f8f9fa",
            fg="#2c3e50",
            width=18
        )
        subject_entry.pack(side=tk.LEFT, padx=(0, 5), ipady=5)
        subject_entry.insert(0, f"Subject {index + 1}")
        self.subject_entries.append(subject_entry)

        marks_entry = tk.Entry(
            row_frame,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bg="#f8f9fa",
            fg="#2c3e50",
            width=18
        )
        marks_entry.pack(side=tk.LEFT, padx=5, ipady=5)
        self.marks_entries.append(marks_entry)

        total_entry = tk.Entry(
            row_frame,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bg="#f8f9fa",
            fg="#2c3e50",
            width=18
        )
        total_entry.pack(side=tk.LEFT, padx=5, ipady=5)
        total_entry.insert(0, "100")
        self.total_marks_entries.append(total_entry)

    def create_buttons_section(self, parent):
        buttons_frame = tk.Frame(parent, bg="#ffffff")
        buttons_frame.pack(fill=tk.X, pady=(0, 15))

        self.create_button(
            buttons_frame,
            "Calculate Result",
            "#4a90e2",
            self.calculate_result
        ).pack(side=tk.LEFT, expand=True, padx=3)

        self.create_button(
            buttons_frame,
            "Clear",
            "#95a5a6",
            self.clear_fields
        ).pack(side=tk.LEFT, expand=True, padx=3)

        self.create_button(
            buttons_frame,
            "Generate PDF",
            "#27ae60",
            self.generate_pdf
        ).pack(side=tk.LEFT, expand=True, padx=3)

        self.create_button(
            buttons_frame,
            "Generate Excel",
            "#f39c12",
            self.generate_excel
        ).pack(side=tk.LEFT, expand=True, padx=3)

    def create_results_section(self, parent):
        results_frame = tk.LabelFrame(
            parent,
            text="Results Summary",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        results_frame.pack(fill=tk.X)

        self.total_label = self.create_result_label(results_frame, "Total Marks:", 0)
        self.percentage_label = self.create_result_label(results_frame, "Percentage:", 1)
        self.grade_label = self.create_result_label(results_frame, "Grade:", 2)
        self.result_label = self.create_result_label(results_frame, "Result:", 3)

    def create_chart_section(self, parent):
        chart_frame = tk.LabelFrame(
            parent,
            text="Performance Chart",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        self.chart_container = tk.Frame(chart_frame, bg="#ffffff")
        self.chart_container.pack(fill=tk.BOTH, expand=True)

    def create_input_field(self, parent, label_text, row):
        frame = tk.Frame(parent, bg="#ffffff")
        frame.pack(fill=tk.X, pady=5)

        label = tk.Label(
            frame,
            text=label_text,
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#2c3e50",
            width=15,
            anchor='w'
        )
        label.pack(side=tk.LEFT)

        entry = tk.Entry(
            frame,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bg="#f8f9fa",
            fg="#2c3e50"
        )
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)

        return entry

    def create_button(self, parent, text, color, command):
        button = tk.Button(
            parent,
            text=text,
            font=("Segoe UI", 10, "bold"),
            bg=color,
            fg="white",
            relief=tk.FLAT,
            cursor="hand2",
            command=command,
            padx=10,
            pady=8
        )

        button.bind("<Enter>", lambda e: button.config(bg=self.lighten_color(color)))
        button.bind("<Leave>", lambda e: button.config(bg=color))

        return button

    def create_result_label(self, parent, label_text, row):
        frame = tk.Frame(parent, bg="#ffffff")
        frame.pack(fill=tk.X, pady=3)

        label = tk.Label(
            frame,
            text=label_text,
            font=("Segoe UI", 10, "bold"),
            bg="#ffffff",
            fg="#2c3e50",
            width=15,
            anchor='w'
        )
        label.pack(side=tk.LEFT)

        value_label = tk.Label(
            frame,
            text="--",
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#4a90e2",
            anchor='w'
        )
        value_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        return value_label

    def lighten_color(self, color):
        colors = {
            "#4a90e2": "#5fa3f5",
            "#95a5a6": "#a8b8b9",
            "#27ae60": "#2ecc71",
            "#f39c12": "#f5a623"
        }
        return colors.get(color, color)

    def validate_inputs(self):
        if not self.name_entry.get().strip():
            messagebox.showerror("Error", "Please enter student name")
            return False

        if not self.roll_entry.get().strip():
            messagebox.showerror("Error", "Please enter roll number")
            return False

        if not self.class_entry.get().strip():
            messagebox.showerror("Error", "Please enter class/section")
            return False

        for i in range(5):
            try:
                marks = float(self.marks_entries[i].get())
                total = float(self.total_marks_entries[i].get())

                if marks < 0 or total < 0:
                    messagebox.showerror("Error", f"Marks cannot be negative for {self.subject_entries[i].get()}")
                    return False

                if marks > total:
                    messagebox.showerror("Error", f"Marks obtained cannot exceed total marks for {self.subject_entries[i].get()}")
                    return False

            except ValueError:
                messagebox.showerror("Error", f"Please enter valid numbers for {self.subject_entries[i].get()}")
                return False

        return True

    def calculate_result(self):
        if not self.validate_inputs():
            return

        total_marks_obtained = 0
        total_marks_possible = 0
        subject_data = []

        for i in range(5):
            marks = float(self.marks_entries[i].get())
            total = float(self.total_marks_entries[i].get())
            subject_name = self.subject_entries[i].get()

            total_marks_obtained += marks
            total_marks_possible += total
            subject_data.append((subject_name, marks, total))

        percentage = (total_marks_obtained / total_marks_possible) * 100
        grade = self.calculate_grade(percentage)
        result = "Pass" if percentage >= 33 else "Fail"

        self.total_label.config(text=f"{total_marks_obtained}/{total_marks_possible}")
        self.percentage_label.config(text=f"{percentage:.2f}%")
        self.grade_label.config(text=grade)
        self.result_label.config(
            text=result,
            fg="#27ae60" if result == "Pass" else "#e74c3c"
        )

        self.display_chart(subject_data)

    def calculate_grade(self, percentage):
        if percentage >= 90:
            return "A+"
        elif percentage >= 80:
            return "A"
        elif percentage >= 70:
            return "B+"
        elif percentage >= 60:
            return "B"
        elif percentage >= 50:
            return "C"
        elif percentage >= 40:
            return "D"
        elif percentage >= 33:
            return "E"
        else:
            return "F"

    def display_chart(self, subject_data):
        for widget in self.chart_container.winfo_children():
            widget.destroy()

        subjects = [s[0] for s in subject_data]
        marks = [s[1] for s in subject_data]
        totals = [s[2] for s in subject_data]

        fig, ax = plt.subplots(figsize=(5, 4), facecolor='white')

        x = range(len(subjects))
        width = 0.35

        bars1 = ax.bar([i - width/2 for i in x], marks, width, label='Marks Obtained', color='#4a90e2')
        bars2 = ax.bar([i + width/2 for i in x], totals, width, label='Total Marks', color='#e0e0e0')

        ax.set_xlabel('Subjects', fontsize=10)
        ax.set_ylabel('Marks', fontsize=10)
        ax.set_title('Subject-wise Performance', fontsize=11, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(subjects, rotation=15, ha='right', fontsize=8)
        ax.legend(fontsize=8)
        ax.grid(axis='y', alpha=0.3)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.chart_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def clear_fields(self):
        self.name_entry.delete(0, tk.END)
        self.roll_entry.delete(0, tk.END)
        self.class_entry.delete(0, tk.END)

        for i in range(5):
            self.subject_entries[i].delete(0, tk.END)
            self.subject_entries[i].insert(0, f"Subject {i + 1}")
            self.marks_entries[i].delete(0, tk.END)
            self.total_marks_entries[i].delete(0, tk.END)
            self.total_marks_entries[i].insert(0, "100")

        self.total_label.config(text="--")
        self.percentage_label.config(text="--")
        self.grade_label.config(text="--")
        self.result_label.config(text="--", fg="#4a90e2")

        for widget in self.chart_container.winfo_children():
            widget.destroy()

    def generate_pdf(self):
        if not self.validate_inputs():
            return

        if self.total_label.cget("text") == "--":
            messagebox.showwarning("Warning", "Please calculate result first")
            return

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 20)

        pdf.cell(0, 20, "MARKSHEET", ln=True, align="C")
        pdf.ln(10)

        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 10, f"Student Name: {self.name_entry.get()}", ln=True)
        pdf.cell(0, 10, f"Roll Number: {self.roll_entry.get()}", ln=True)
        pdf.cell(0, 10, f"Class/Section: {self.class_entry.get()}", ln=True)
        pdf.cell(0, 10, f"Date: {datetime.now().strftime('%d-%m-%Y')}", ln=True)
        pdf.ln(10)

        pdf.set_font("Arial", "B", 12)
        pdf.cell(70, 10, "Subject", border=1)
        pdf.cell(60, 10, "Marks Obtained", border=1)
        pdf.cell(60, 10, "Total Marks", border=1)
        pdf.ln()

        pdf.set_font("Arial", "", 12)
        for i in range(5):
            pdf.cell(70, 10, self.subject_entries[i].get(), border=1)
            pdf.cell(60, 10, self.marks_entries[i].get(), border=1)
            pdf.cell(60, 10, self.total_marks_entries[i].get(), border=1)
            pdf.ln()

        pdf.ln(10)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, f"Total Marks: {self.total_label.cget('text')}", ln=True)
        pdf.cell(0, 10, f"Percentage: {self.percentage_label.cget('text')}", ln=True)
        pdf.cell(0, 10, f"Grade: {self.grade_label.cget('text')}", ln=True)
        pdf.cell(0, 10, f"Result: {self.result_label.cget('text')}", ln=True)

        filename = f"marksheet_{self.roll_entry.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf.output(filename)

        messagebox.showinfo("Success", f"PDF generated successfully: {filename}")

    def generate_excel(self):
        if not self.validate_inputs():
            return

        if self.total_label.cget("text") == "--":
            messagebox.showwarning("Warning", "Please calculate result first")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marksheet"

        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "MARKSHEET"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws["A3"] = "Student Name:"
        ws["B3"] = self.name_entry.get()
        ws["A4"] = "Roll Number:"
        ws["B4"] = self.roll_entry.get()
        ws["A5"] = "Class/Section:"
        ws["B5"] = self.class_entry.get()
        ws["A6"] = "Date:"
        ws["B6"] = datetime.now().strftime('%d-%m-%Y')

        headers = ["Subject", "Marks Obtained", "Total Marks", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in range(5):
            row = 9 + i
            ws.cell(row, 1).value = self.subject_entries[i].get()
            marks = float(self.marks_entries[i].get())
            total = float(self.total_marks_entries[i].get())
            ws.cell(row, 2).value = marks
            ws.cell(row, 3).value = total
            ws.cell(row, 4).value = f"{(marks/total)*100:.2f}%"

        summary_row = 15
        ws[f"A{summary_row}"] = "Total Marks:"
        ws[f"B{summary_row}"] = self.total_label.cget("text")
        ws[f"A{summary_row+1}"] = "Percentage:"
        ws[f"B{summary_row+1}"] = self.percentage_label.cget("text")
        ws[f"A{summary_row+2}"] = "Grade:"
        ws[f"B{summary_row+2}"] = self.grade_label.cget("text")
        ws[f"A{summary_row+3}"] = "Result:"
        ws[f"B{summary_row+3}"] = self.result_label.cget("text")

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 20

        filename = f"marksheet_{self.roll_entry.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        messagebox.showinfo("Success", f"Excel file generated successfully: {filename}")


def main():
    root = tk.Tk()
    app = MarksheetGenerator(root)
    root.mainloop()


if __name__ == "__main__":
    main()
